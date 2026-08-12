#!/usr/bin/env python3
from datetime import datetime
import json
import os
from pathlib import Path
import openpyxl


def load_dotenv(dotenv_path: str = ".env") -> None:
    """Carrega as variáveis de ambiente do arquivo .env se existir."""
    if not os.path.exists(dotenv_path):
        return

    with open(dotenv_path, encoding="utf-8") as dotenv_file:
        for line in dotenv_file:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue

            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


def resolver_caminho_planilha(planilha_env: str) -> Path | None:
    """
    Resolve o caminho exato do arquivo .xlsx.
    Trata casos onde a variável aponta para o diretório ou diretamente para o arquivo.
    """
    caminho = Path(planilha_env)

    if caminho.is_file() and caminho.suffix.lower() == ".xlsx":
        return caminho

    if caminho.is_dir():
        # Busca o arquivo Planilha_Mestra.xlsx ou qualquer .xlsx dentro do diretório
        arquivo_padrao = caminho / "Planilha_Mestra.xlsx"
        if arquivo_padrao.is_file():
            return arquivo_padrao

        arquivos_xlsx = list(caminho.glob("*.xlsx"))
        if arquivos_xlsx:
            return arquivos_xlsx[0]

    return None


def preencher_planilha(json_file_path: Path, excel_file_path: Path) -> None:
    """Lê o arquivo data.json e insere os dados na planilha Excel preservando cabeçalhos."""
    if not json_file_path.is_file():
        print(f"[ERRO] Arquivo JSON não encontrado em: {json_file_path}")
        return

    with open(json_file_path, "r", encoding="utf-8") as f:
        cadastros = json.load(f)

    if not isinstance(cadastros, list):
        print("[ERRO] O arquivo JSON precisa conter uma lista de objetos.")
        return

    # Carrega a planilha existente
    wb = openpyxl.load_workbook(str(excel_file_path))
    sheet = wb.active  # Pega a aba ativa (Planilha_Mestra)

    # Identifica a última linha preenchida com dados reais
    proxima_linha = 2
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(row=r, column=1).value is not None:
            proxima_linha = r + 1

    data_hoje = datetime.now().strftime("%d/%m/%Y")
    registros_inseridos = 0

    for registro in cadastros:
        sheet.cell(row=proxima_linha, column=1, value=registro.get("cpf"))
        sheet.cell(row=proxima_linha, column=2, value=registro.get("nome_completo"))
        sheet.cell(row=proxima_linha, column=3, value=registro.get("data_nascimento"))
        sheet.cell(row=proxima_linha, column=4, value=registro.get("endereco_completo"))
        sheet.cell(row=proxima_linha, column=5, value=registro.get("email"))
        sheet.cell(row=proxima_linha, column=6, value=registro.get("telefone_whatsapp"))
        sheet.cell(row=proxima_linha, column=7, value="PROCESSADO")
        sheet.cell(row=proxima_linha, column=8, value=data_hoje)
        sheet.cell(row=proxima_linha, column=9, value="Dados extraídos via OCR/PDF")

        proxima_linha += 1
        registros_inseridos += 1

    wb.save(str(excel_file_path))
    print(f"[OK] {registros_inseridos} registros inseridos com sucesso na planilha {excel_file_path.name}")


def main() -> int:
    load_dotenv()

    json_data_dir = os.getenv("JSON_DATA_DIR", "json_data")
    planilha_mestra_env = os.getenv("PLANILHA_MESTRA_DIR", ".")

    json_file_path = Path(json_data_dir) / "data.json"
    excel_file_path = resolver_caminho_planilha(planilha_mestra_env)

    if not excel_file_path:
        print(f"[ERRO] Nenhum arquivo .xlsx encontrado no caminho: {planilha_mestra_env}")
        return 1

    print(f"Lendo dados de: {json_file_path}")
    print(f"Atualizando planilha: {excel_file_path}\n")

    preencher_planilha(json_file_path, excel_file_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())