from playwright.sync_api import sync_playwright, TimeoutError
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
import time
import re
import math
import csv
import pandas as pd

def calcular_correios(page, origem, destino, altura, largura, comprimento):
    page.goto("https://www.correios.com.br/")
    page.wait_for_load_state('domcontentloaded')

    page.fill('#origem', origem)
    page.fill('#destino', destino)
    page.fill('#altura', formatar_dimensao_arredondar(altura))
    page.fill('#largura', formatar_dimensao_arredondar(largura))
    page.fill('#comprimento', formatar_dimensao_arredondar(comprimento))

    time.sleep(1)

    page.click('[data-botao-simular="simular"]', force=True)

    page.wait_for_selector(
        '[data-respostas="todas-respostas"], [data-paragrafo-erro="mensagemparagrafoErro"]'
    )

    erro_elem = page.query_selector('[data-paragrafo-erro="mensagemparagrafoErro"]')
    if erro_elem:
        return {"erro": erro_elem.text_content()}

    produtos = page.query_selector_all('[data-resposta-produto="produto"]')

    resultado = {}

    for produto in produtos:
        nome = produto.query_selector("img").get_attribute("alt")
        preco = produto.query_selector('[data-mensagem-preco="mensagemPrecoPPN"] span').text_content().replace("*", "").strip()
        prazo = produto.query_selector('[data-mensagem-prazo="mensagemPrazo"] span').text_content()

        if "Sedex" in nome:
            resultado["SEDEX"] = {"preco": preco, "prazo": prazo}
        elif "PAC" in nome:
            resultado["PAC"] = {"preco": preco, "prazo": prazo}

    return resultado


def calcular_loggi(page, origem, destino, altura, largura, comprimento, peso, valor):
    page.goto("https://www.loggi.com/")
    page.wait_for_load_state('domcontentloaded')

    try:
        page.locator('[data-testid="close-button"]').click(timeout=3000)
    except:
        pass

    page.locator("text=Calcular frete").click()

    page.locator('[placeholder="CEP de origem"]').fill(origem)
    page.locator('[placeholder="CEP de destino"]').fill(destino)
    page.locator('input[name="height"]').fill(formatar_dimensao_arredondar(altura))
    page.locator('input[name="width"]').fill(formatar_dimensao_arredondar(largura))
    page.locator('input[name="length"]').fill(formatar_dimensao_arredondar(comprimento))
    page.locator('input[name="weight"]').fill(formatar_peso_loggi(peso))
    page.locator('input[name="itemValue"]').fill(formatar_peso_loggi(valor))

    if page.locator("text=Aceitar").is_visible():
        page.locator("text=Aceitar").click()

    #time.sleep(5)

    simular_btn = page.locator('text=Simular envio')
    page.wait_for_selector('text=Simular envio')

    if not simular_btn.is_enabled():
        mensagens = page.locator('p[id$="-form-item-message"].text-v2-feedback-negative')
        if mensagens.count() > 0:
            for i in range(mensagens.count()):
                return("erro:", mensagens.nth(i).inner_text())
        else:        
            return {"erro": "Dados inválidos"}

    simular_btn.click()

    try:
        page.wait_for_function("""
            () => {
                return document.body.innerText.includes('Área não atendida') ||
                    document.body.innerText.includes('Falha ao realizar cotação');
            }
        """, timeout=3000)

        if page.locator('text=Área não atendida').is_visible():
            return {"erro": "Área não atendida"}

        if page.locator('text=Falha ao realizar cotação').is_visible():
            return {"erro": "Falha ao realizar cotação"}

    except TimeoutError:
        pass

    valores = page.locator('h2:has-text("R$")')
    prazos = page.locator('span').filter(has_text="Chega em até")
    tipos = page.locator('span').filter(has_text="Loggi Entrega")

    page.wait_for_selector('h2:has-text("R$")')

    resultado = []

    for i in range(tipos.count()):
        resultado.append({
            "tipo": tipos.nth(i).inner_text(),
            "valor": valores.nth(i).inner_text(),
            "prazo": prazos.nth(i).inner_text()
        })

    return resultado


def calcular_jt(page, origem, destino, altura, largura, comprimento, peso, valor):
    page.goto("https://www.jtexpress.com.br/")
    page.wait_for_load_state('domcontentloaded')

    page.locator('input[placeholder="00000-000"]').nth(0).fill(origem)
    page.locator('input[placeholder="00000-000"]').nth(1).fill(destino)
    page.locator('input[placeholder="0,0"]').nth(0).fill(formatar_dimensao_jt(altura))
    page.locator('input[placeholder="0,0"]').nth(1).fill(formatar_dimensao_jt(largura))
    page.locator('input[placeholder="0,0"]').nth(2).fill(formatar_dimensao_jt(comprimento))
    page.locator('input[placeholder="0,000"]').fill(formatar_peso_jt(peso))
    page.locator('input[maxlength="7"]').fill(formatar_valor_jt(valor))

    #time.sleep(5)

    page.locator('button:has-text("Calcular Frete")').click()

    page.wait_for_function("""
        () => {
            const temResultado = Array.from(document.querySelectorAll('span'))
                .some(el => el.innerText.includes('R$') || el.innerText.includes('Entregue em'));
            const temErro = document.querySelectorAll('.el-form-item__error').length > 0;
            return temResultado || temErro;
        }
    """)

    erros = page.locator('.el-form-item__error')
    if erros.count() > 0:
        return {"erro": [erros.nth(i).inner_text() for i in range(erros.count())]}

    valor_res = page.locator('span').filter(has_text="R$").first.inner_text()
    prazo_res = page.locator('span').filter(has_text="Entregue em").first.inner_text()

    return {"valor": valor_res, "prazo": prazo_res}


def calcular_todos(origem, destino, altura, largura, comprimento, peso, valor):
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            args=["--start-maximized"]
        )

        context = browser.new_context(no_viewport=True)

        resultados = {}

        # Correios
        page = context.new_page()
        resultados["correios"] = calcular_correios(
            page, origem, destino, altura, largura, comprimento
        )
        page.close()

        # Loggi
        page = context.new_page()
        resultados["loggi"] = calcular_loggi(
            page, origem, destino, altura, largura, comprimento, peso, valor
        )
        page.close()

        # J&T
        page = context.new_page()
        resultados["jt"] = calcular_jt(
            page, origem, destino, altura, largura , comprimento, peso, valor
        )
        page.close()

        context.close()

        return resultados

def limpar_cep(cep):
    cep = str(cep)
    cep_limpo = re.sub(r'\D', '', cep)

    if len(cep_limpo) != 8:
        raise ValueError(f"CEP inválido: {cep}")

    return cep_limpo

def formatar_dimensao_arredondar(valor):
    return str(math.ceil(float(str(valor).replace(",", "."))))

def formatar_peso_loggi(valor):
    valor_float = float(str(valor).replace(",", "."))
    return str(valor_float).replace(".", ",")

def formatar_dimensao_jt(valor):
    valor_str = str(valor).replace(",", ".").strip()

    if "." in valor_str:
        partes = valor_str.split(".")
        resultado = partes[0] + partes[1]  
    else:
        resultado = valor_str + "0"

    resultado = re.sub(r"\D", "", resultado)
    return resultado

def formatar_peso_jt(valor):
    valor_str = str(valor).replace(",", ".").strip()

    if "." in valor_str:
        partes = valor_str.split(".")
        resultado = partes[0] + partes[1] + "0" 
    else:
        resultado = valor_str + "000"

    resultado = re.sub(r"\D", "", resultado)
    return resultado

def formatar_valor_jt(valor):
    valor_str = str(valor).replace(",", ".").strip()

    if "." in valor_str:
        partes = valor_str.split(".")
        resultado = partes[0] + partes[1] + "0"
    else:
        resultado = valor_str + "00"

    resultado = re.sub(r"\D", "", resultado)
    return resultado

def processar_transportadora(nome_arquivo_csv, transportadora):
    pedidos = []
    with open(nome_arquivo_csv, newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            pedidos.append(row)

    resultados = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        for pedido in pedidos:
            origem = limpar_cep(pedido['cep_origem'])
            destino = limpar_cep(pedido['cep_destino'])
            altura = float(pedido['altura'])
            largura = float(pedido['largura'])
            comprimento = float(pedido['comprimento'])
            peso = float(pedido['peso'])
            valor = float(pedido['valor'])

            if transportadora == "correios":
                res = calcular_correios(page, origem, destino, altura, largura, comprimento)
            elif transportadora == "loggi":
                res = calcular_loggi(page, origem, destino, altura, largura, comprimento, peso, valor)
            elif transportadora == "jt":
                res = calcular_jt(page, origem, destino, altura, largura, comprimento, peso, valor)
            else:
                res = {"erro": "Transportadora inválida"}

            resultados.append({
                "cliente": pedido["cliente"],
                "produto": pedido["produto"],
                "transportadora": transportadora,
                "resultado": res
            })

        browser.close()

    return resultados

def mesclar_centralizar_e_pintar(excel_file, cor_fundo="FF0000"):
    wb = load_workbook(excel_file)
    ws = wb.active

    fill = PatternFill(start_color=cor_fundo, end_color=cor_fundo, fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        start_col = None
        last_val = None
        for col_idx, cell in enumerate(row, 1):
            if cell.value == last_val and last_val is not None:
                # continua sequência
                end_col = col_idx
            else:
                # finalizar sequência anterior
                if last_val is not None and start_col is not None and end_col > start_col:
                    ws.merge_cells(start_row=cell.row, start_column=start_col,
                                   end_row=cell.row, end_column=end_col)
                    merged_cell = ws.cell(row=cell.row, column=start_col)
                    merged_cell.alignment = alignment
                    merged_cell.fill = fill
                # iniciar nova sequência
                start_col = col_idx
                end_col = col_idx
                last_val = cell.value
        # mesclar última sequência da linha
        if last_val is not None and start_col is not None and end_col > start_col:
            ws.merge_cells(start_row=row[0].row, start_column=start_col,
                           end_row=row[0].row, end_column=end_col)
            merged_cell = ws.cell(row=row[0].row, column=start_col)
            merged_cell.alignment = alignment
            merged_cell.fill = fill

    wb.save(excel_file)
    print("Mesclagem, centralização e pintura concluídas!")

def gerar_html_comparativo(excel_file="comparativo_fretes.xlsx", html_file="comparativo_fretes.html"):
    df = pd.read_excel(excel_file, header=[0, 1, 2], index_col=0)
    
    # Substituir NaN por vazio
    df = df.fillna("")
    
    # Remover "Unnamed" dos cabeçalhos
    df.columns = pd.MultiIndex.from_tuples([
        tuple("" if "Unnamed" in str(col) else col for col in cols)
        for cols in df.columns
    ])
    
    # Estilo moderno com fundo escuro
    styles = [
        # Estilo global da tabela
        dict(selector="table", props=[
            ("width", "100%"),
            ("border-collapse", "separate"),
            ("border-spacing", "0"),
            ("background-color", "#1a1a1a"),
            ("border-radius", "12px"),
            ("overflow", "hidden"),
            ("box-shadow", "0 4px 6px rgba(0,0,0,0.3)")
        ]),
        
        # Estilo do cabeçalho principal
        dict(selector="th", props=[
            ("background-color", "#2d2d2d"),
            ("color", "#ffffff"),
            ("text-align", "center"),
            ("padding", "12px 15px"),
            ("font-weight", "600"),
            ("font-size", "14px"),
            ("border", "none"),
            ("border-bottom", "2px solid #3b8ed0")
        ]),
        
        # Estilo das células
        dict(selector="td", props=[
            ("text-align", "center"),
            ("padding", "10px 12px"),
            ("border", "none"),
            ("border-bottom", "1px solid #333333"),
            ("color", "#e0e0e0"),
            ("font-size", "13px"),
            ("background-color", "#1e1e1e")
        ]),
        
        # Efeito hover nas linhas
        dict(selector="tr:hover td", props=[
            ("background-color", "#2a2a2a"),
            ("transition", "background-color 0.3s ease")
        ]),
        
        # Estilo para células de erro
        dict(selector="td.error", props=[
            ("color", "#e06c75"),
            ("font-weight", "500")
        ]),
        
        # Estilo para células de sucesso/valor
        dict(selector="td.valor", props=[
            ("color", "#98c379"),
            ("font-weight", "600")
        ]),
        
        # Estilo para o primeiro nível do cabeçalho
        dict(selector="thead tr:first-child th", props=[
            ("background-color", "#252525"),
            ("font-size", "15px"),
            ("padding", "15px"),
            ("letter-spacing", "0.5px")
        ]),
        
        # Estilo para o segundo nível do cabeçalho
        dict(selector="thead tr:last-child th", props=[
            ("background-color", "#2d2d2d"),
            ("font-size", "13px"),
            ("font-weight", "500"),
            ("color", "#b0b0b0")
        ]),
        
        # Estilo para células vazias
        dict(selector="td:empty", props=[
            ("background-color", "#181818"),
            ("color", "#666666")
        ]),
        
        # Estilo para a primeira coluna (índice)
        dict(selector="td:first-child, th:first-child", props=[
            ("background-color", "#252525"),
            ("font-weight", "600"),
            ("border-right", "1px solid #333333")
        ]),
        
        # Animação suave
        dict(selector="*", props=[
            ("transition", "all 0.2s ease")
        ])
    ]
    
    # Aplicar estilo condicional às células
    def highlight_cells(val):
        """Aplica estilo especial para valores específicos"""
        if isinstance(val, str):
            if "erro" in val.lower() or "não atendida" in val.lower():
                return 'class="error"'
            if "R$" in val or val.replace(".", "").replace(",", "").isdigit():
                return 'class="valor"'
        return ''
    
    # Converter para HTML com estilo
    html = df.style.set_table_styles(styles).set_table_attributes('class="dataframe"').to_html()
    
    # Adicionar CSS e JavaScript moderno
    full_html = f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Comparativo de Fretes - Dashboard</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            background: linear-gradient(135deg, #1a1a1a 0%, #0d0d0d 100%);
            padding: 20px;
            min-height: 100vh;
        }}
        
        .container {{
            margin: 0 auto;
            background: #1a1a1a;
            border-radius: 16px;
            overflow: hidden;
            box-shadow: 0 20px 60px rgba(0,0,0,0.5);
        }}
        
        .header {{
            background: linear-gradient(135deg, #2d2d2d 0%, #1a1a1a 100%);
            padding: 30px;
            border-bottom: 1px solid #3b8ed0;
        }}
        
        .header h1 {{
            color: #ffffff;
            font-size: 28px;
            font-weight: 600;
            margin-bottom: 10px;
            display: flex;
            align-items: center;
            gap: 10px;
        }}
        
        .header p {{
            color: #b0b0b0;
            font-size: 14px;
        }}
        
        .stats {{
            display: flex;
            gap: 20px;
            padding: 20px 30px;
            background: #1e1e1e;
            border-bottom: 1px solid #333;
        }}
        
        .stat-card {{
            flex: 1;
            background: #2d2d2d;
            padding: 15px;
            border-radius: 10px;
            text-align: center;
            transition: transform 0.2s;
        }}
        
        .stat-card:hover {{
            transform: translateY(-2px);
        }}
        
        .stat-card .label {{
            color: #b0b0b0;
            font-size: 12px;
            text-transform: uppercase;
            letter-spacing: 1px;
            margin-bottom: 5px;
        }}
        
        .stat-card .value {{
            color: #3b8ed0;
            font-size: 24px;
            font-weight: 600;
        }}
        
        .table-wrapper {{
            overflow-x: auto;
            padding: 20px;
            background: #1a1a1a;
        }}
        
        .footer {{
            background: #1e1e1e;
            padding: 20px;
            text-align: center;
            border-top: 1px solid #333;
            color: #888;
            font-size: 12px;
        }}
        
        @keyframes fadeIn {{
            from {{
                opacity: 0;
                transform: translateY(20px);
            }}
            to {{
                opacity: 1;
                transform: translateY(0);
            }}
        }}
        
        .container {{
            animation: fadeIn 0.5s ease-out;
        }}
        
        /* Scrollbar personalizada */
        ::-webkit-scrollbar {{
            width: 10px;
            height: 10px;
        }}
        
        ::-webkit-scrollbar-track {{
            background: #1a1a1a;
        }}
        
        ::-webkit-scrollbar-thumb {{
            background: #3b8ed0;
            border-radius: 5px;
        }}
        
        ::-webkit-scrollbar-thumb:hover {{
            background: #2c6ea8;
        }}
        
        /* Estilos para badges */
        .badge {{
            display: inline-block;
            padding: 4px 8px;
            border-radius: 6px;
            font-size: 11px;
            font-weight: 600;
            margin-left: 8px;
        }}
        
        .badge-success {{
            background: rgba(152, 195, 121, 0.2);
            color: #98c379;
        }}
        
        .badge-error {{
            background: rgba(224, 108, 117, 0.2);
            color: #e06c75;
        }}
        
        /* Responsividade */
        @media (max-width: 768px) {{
            .stats {{
                flex-direction: column;
            }}
            
            .header h1 {{
                font-size: 22px;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>
                🚚 Comparativo de Fretes
            </h1>
            <p>Análise comparativa entre Correios, Loggi e J&T Express</p>
        </div>
        
        <div class="table-wrapper">
            {html}
        </div>
    </div>
    
    <script>
        // Adicionar tooltips nas células
        document.querySelectorAll('td').forEach(cell => {{
            if (cell.textContent.includes('erro') || cell.textContent.includes('não atendida')) {{
                cell.style.cursor = 'help';
                cell.title = 'Falha na cotação para esta transportadora';
            }}
            if (cell.textContent.includes('R$')) {{
                cell.style.fontWeight = 'bold';
            }}
        }});
        
        // Adicionar contador de pedidos
        const rows = document.querySelectorAll('tbody tr');
        const statCards = document.querySelectorAll('.stat-card');
        if (statCards.length > 0 && rows.length > 0) {{
            const pedidosCount = document.createElement('div');
            pedidosCount.className = 'stat-card';
            pedidosCount.innerHTML = `
                <div class="label">Pedidos Analisados</div>
                <div class="value">$\{{rows.length}}</div>
            `;
            document.querySelector('.stats').appendChild(pedidosCount);
        }}
        
        // Efeito de brilho nas células de valor
        document.querySelectorAll('td.valor, td:contains("R$")').forEach(cell => {{
            cell.style.position = 'relative';
            cell.style.overflow = 'hidden';
        }});
        
        // Função para encontrar células com R$
        const cellsWithMoney = Array.from(document.querySelectorAll('td')).filter(cell => 
            cell.textContent.includes('R$')
        );
        cellsWithMoney.forEach(cell => {{
            cell.classList.add('valor');
        }});
    </script>
</body>
</html>"""
    
    # Salvar o arquivo
    with open(html_file, "w", encoding="utf-8") as f:
        f.write(full_html)
    
    print(f"✨ Arquivo HTML moderno gerado com sucesso: {html_file}")
    return html_file

def gerar_planilha_final(resultados_correios, resultados_loggi, resultados_jt):
    linhas = []

    for i in range(len(resultados_correios)):
        correios_res = resultados_correios[i]["resultado"]
        loggi_res = resultados_loggi[i]["resultado"]
        jt_res = resultados_jt[i]["resultado"]

        linha = {
            # Cliente e Produto
            ("", "", "Cliente"): resultados_correios[i]["cliente"],
            ("", "", "Produto"): resultados_correios[i]["produto"],
        }

        # --- Correios ---
        correios_res = resultados_correios[i]["resultado"]
        if "erro" in correios_res:
            linha.update({
                ("Correios", "Sedex", "Valor"): correios_res["erro"],
                ("Correios", "Sedex", "Prazo"): correios_res["erro"],
                ("Correios", "PAC", "Valor"): correios_res["erro"],
                ("Correios", "PAC", "Prazo"): correios_res["erro"],
            })
        else:
            linha.update({
                ("Correios", "Sedex", "Valor"): correios_res.get("SEDEX", {}).get("preco", ""),
                ("Correios", "Sedex", "Prazo"): correios_res.get("SEDEX", {}).get("prazo", ""),
                ("Correios", "PAC", "Valor"): correios_res.get("PAC", {}).get("preco", ""),
                ("Correios", "PAC", "Prazo"): correios_res.get("PAC", {}).get("prazo", ""),
            })

        # --- Loggi ---
        loggi_res = resultados_loggi[i]["resultado"]
        if isinstance(loggi_res, dict) and "erro" in loggi_res:
            linha.update({
                ("Loggi", "Entrega Local", "Valor"): loggi_res["erro"],
                ("Loggi", "Entrega Local", "Prazo"): loggi_res["erro"],
                ("Loggi", "Entrega Nacional", "Valor"): loggi_res["erro"],
                ("Loggi", "Entrega Nacional", "Prazo"): loggi_res["erro"],
            })
        else:
            for opcao in loggi_res:
                tipo = opcao.get("tipo", "").lower()
                if "local" in tipo:
                    linha[("Loggi", "Entrega Local", "Valor")] = opcao.get("valor", "")
                    linha[("Loggi", "Entrega Local", "Prazo")] = opcao.get("prazo", "")
                elif "nacional" in tipo:
                    linha[("Loggi", "Entrega Nacional", "Valor")] = opcao.get("valor", "")
                    linha[("Loggi", "Entrega Nacional", "Prazo")] = opcao.get("prazo", "")

        # --- J&T ---
        jt_res = resultados_jt[i]["resultado"]
        if "erro" in jt_res:
            linha.update({
                ("J&T", "", "Valor"): jt_res["erro"],
                ("J&T", "", "Prazo"): jt_res["erro"],
            })
        else:
            linha.update({
                ("J&T", "", "Valor"): jt_res.get("valor", ""),
                ("J&T", "", "Prazo"): jt_res.get("prazo", ""),
            })

        # Preencher dados do Loggi
        if isinstance(loggi_res, list):
            for opcao in loggi_res:
                tipo = opcao.get("tipo", "").lower()

                if "local" in tipo:
                    linha[("Loggi", "Entrega Local", "Valor")] = opcao.get("valor", "")
                    linha[("Loggi", "Entrega Local", "Prazo")] = opcao.get("prazo", "")

                elif "nacional" in tipo:
                    linha[("Loggi", "Entrega Nacional", "Valor")] = opcao.get("valor", "")
                    linha[("Loggi", "Entrega Nacional", "Prazo")] = opcao.get("prazo", "")

        linhas.append(linha)

    # Criar DataFrame com MultiIndex
    df = pd.DataFrame(linhas)

    # Definir a ordem exata das colunas
    colunas_ordenadas = [
        ("", "", "Cliente"),
        ("", "", "Produto"),
        ("Correios", "Sedex", "Valor"),
        ("Correios", "Sedex", "Prazo"),
        ("Correios", "PAC", "Valor"),
        ("Correios", "PAC", "Prazo"),
        ("Loggi", "Entrega Local", "Valor"),
        ("Loggi", "Entrega Local", "Prazo"),
        ("Loggi", "Entrega Nacional", "Valor"),
        ("Loggi", "Entrega Nacional", "Prazo"),
        ("J&T", "", "Valor"),
        ("J&T", "", "Prazo"),
    ]

    df = df[colunas_ordenadas]

    # Converter colunas para MultiIndex
    df.columns = pd.MultiIndex.from_tuples(df.columns)

    # Exportar para Excel com cabeçalho hierárquico
    df.to_excel("comparativo_fretes.xlsx", index=True, merge_cells=True)
    mesclar_centralizar_e_pintar("comparativo_fretes.xlsx")
    gerar_html_comparativo("comparativo_fretes.xlsx", "comparativo_fretes.html")

    print("Planilha comparativa gerada com sucesso!")

if __name__ == "__main__":
    csv_input = "pedidos_valor.csv"  

    resultados_correios = processar_transportadora(csv_input, "correios")
    print("Correios concluído!")
    print(resultados_correios)

    resultados_loggi = processar_transportadora(csv_input, "loggi")
    print("Loggi concluído!")
    print(resultados_loggi)

    resultados_jt = processar_transportadora(csv_input, "jt")
    print("J&T concluído!")
    print(resultados_jt)

    gerar_planilha_final(resultados_correios, resultados_loggi, resultados_jt)
